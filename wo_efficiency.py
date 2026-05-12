
#"""wo_efficiency.py (fast + cache-safe)

#Fixes two real-world issues:
#1) Performance: Parses IndySoft "All WOs" export using openpyxl iter_rows(values_only=True),
   #avoids per-cell access, and adds early-stop for long blank tails.
#2)
# #In particular, WO_Techs_CompleteOnly is returned as a pipe-delimited string.

#Expected files:
#- All WOs.xlsx (export) with sheet "All WOs"
#- Names_Numbers.xlsx with columns: "Employee #" and "Tech Name"

#All Events expectations (from All Data.xlsx):
#- Columns: Work Order, I.D., Event Type, Date, Time, Total Time
#- Receiving timestamp is Event Type == "Receiving In-Shop"

from __future__ import annotations

from datetime import datetime, date, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook


# -----------------------------
# Tech mapping
# -----------------------------

def load_tech_map(names_numbers_path: str | Path) -> Dict[str, str]:
    df = pd.read_excel(names_numbers_path, engine='openpyxl')
    df = df.rename(columns=lambda c: str(c).strip())
    if 'Employee #' not in df.columns or 'Tech Name' not in df.columns:
        raise ValueError("Names_Numbers.xlsx must contain 'Employee #' and 'Tech Name' columns")
    df['Employee #'] = df['Employee #'].astype(str).str.strip()
    df['Tech Name'] = df['Tech Name'].astype(str).str.strip()
    return dict(zip(df['Employee #'], df['Tech Name']))


def _combine_date_time(d: Any, t: Any) -> Optional[datetime]:
    if d is None or d == '':
        return None

    if isinstance(d, datetime):
        d0 = d.date()
    elif isinstance(d, date):
        d0 = d
    else:
        try:
            d0 = pd.to_datetime(d).date()
        except Exception:
            return None

    if t is None or t == '':
        return datetime.combine(d0, time(0, 0, 0))

    if isinstance(t, datetime):
        t0 = t.time()
    elif isinstance(t, time):
        t0 = t
    else:
        try:
            t0 = pd.to_datetime(t).time()
        except Exception:
            t0 = time(0, 0, 0)

    return datetime.combine(d0, t0)


# -----------------------------
# 1) Parse sub-tables from All WOs export (FAST)
# -----------------------------

def parse_all_wos_subtables(
    all_wos_path: str | Path,
    names_numbers_path: str | Path,
    sheet_name: str = 'All WOs',
    max_scan_rows_for_main_header: int = 80,
    max_cols_scan: int = 220,
    blank_tail_stop: int = 800,
) -> pd.DataFrame:
    """Flatten IndySoft All WOs export into one row per equipment calibration line.

    Performance notes:
    - Uses ws.iter_rows(values_only=True) for speed.
    - Stops early after blank_tail_stop consecutive fully-blank rows.
    """

    tech_map = load_tech_map(names_numbers_path)

    wb = load_workbook(all_wos_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {all_wos_path}. Found: {wb.sheetnames}")
    ws = wb[sheet_name]

    # 1a) Locate main header row containing 'Work Order'
    header_row_idx = None
    header_values: Optional[Tuple[Any, ...]] = None

    for r_i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows_for_main_header, max_col=80, values_only=True),
        start=1,
    ):
        if row and any(v == 'Work Order' for v in row):
            header_row_idx = r_i
            header_values = row
            break

    if header_row_idx is None or header_values is None:
        raise ValueError("Could not find main header row containing 'Work Order'.")

    # Map main header name -> 0-based index
    main_col_map: Dict[str, int] = {}
    for idx0, v in enumerate(header_values):
        if v is None:
            continue
        name = str(v).strip()
        if name:
            main_col_map[name] = idx0

    work_order_idx = main_col_map.get('Work Order')
    if work_order_idx is None:
        raise ValueError("Main header row found but missing 'Work Order' column.")

    # 1b) Iterate subsequent rows
    rows_out: List[Dict[str, Any]] = []
    current_work_order: Optional[str] = None
    sub_col_map: Optional[Dict[str, int]] = None  # 0-based index
    in_subtable = False
    blank_streak = 0

    # Pre-slice check width for "blank" detection
    blank_check_cols = min(30, max_cols_scan)

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_col=max_cols_scan, values_only=True):
        # Normalize row to length max_cols_scan
        if row is None:
            row = tuple()
        if len(row) < max_cols_scan:
            row = tuple(row) + (None,) * (max_cols_scan - len(row))

        # Early stop on long blank tail
        if all((row[c] is None or row[c] == '') for c in range(blank_check_cols)):
            blank_streak += 1
            if blank_streak >= blank_tail_stop:
                break
        else:
            blank_streak = 0

        # Detect new WO block
        wo_val = row[work_order_idx] if work_order_idx < len(row) else None
        if wo_val not in (None, '', 'Work Order'):
            current_work_order = str(wo_val).strip()
            in_subtable = False
            sub_col_map = None
            continue

        # Detect sub-table header signature
        # Export shows: col2 '#', col3 'Related Event Type'
        if len(row) >= 3 and row[1] == '#' and str(row[2]).strip() == 'Related Event Type':
            sub_col_map = {}
            for idx0, v in enumerate(row):
                if v is None:
                    continue
                name = str(v).strip()
                if name:
                    sub_col_map[name] = idx0
            in_subtable = True
            continue

        # Parse sub-table rows
        if in_subtable and sub_col_map and current_work_order:
            idx_col = sub_col_map.get('#', 1)
            idx_val = row[idx_col] if idx_col < len(row) else None

            if idx_val in (None, ''):
                in_subtable = False
                sub_col_map = None
                continue

            def get(col_name: str) -> Any:
                c0 = sub_col_map.get(col_name)
                if c0 is None:
                    return None
                return row[c0] if c0 < len(row) else None

            performed_id = get('Performed By')
            performed_id = None if performed_id in (None, '') else str(performed_id).strip()
            performed_name = tech_map.get(performed_id) if performed_id else None

            completed_dt = _combine_date_time(get('Completed Date'), get('Completed Time'))

            rows_out.append({
                'Work Order': current_work_order,
                'Line #': idx_val,
                'Related Event Type': get('Related Event Type'),
                'Status': get('Status'),
                'CompletedDateTime': completed_dt,
                'PerformedBy_ID': performed_id,
                'PerformedBy_Name': performed_name,
                'Company': get('Company'),
                'Equipment_ID': get('I.D.'),
                'Description': get('Description'),
                'Notes': get('Notes'),
                'Est_Time': get('Est. Time'),
                'Actual_Time': get('Actual Time'),
                'Tracking Status': get('Tracking Status'),
                'Template Name': get('Template Name'),
                'Manufacturer': get('Manufacturer'),
                'Model Number': get('Model Number'),
                'Location': get('Location'),
                'Department': get('Department'),
            })

    df = pd.DataFrame(rows_out)

    # Normalize strings
    for col in ['Work Order', 'Equipment_ID', 'Related Event Type', 'Status', 'PerformedBy_Name', 'Tracking Status']:
        if col in df.columns:
            df[col] = df[col].astype('string').str.strip()

    return df


# -----------------------------
# 2) Receiving DateTime per Work Order (and Customer)
# -----------------------------

def _combine_events_date_time(events_df: pd.DataFrame, date_col: str = 'Date', time_col: str = 'Time') -> pd.Series:
    return pd.to_datetime(
        events_df[date_col].astype(str).str.strip() + ' ' + events_df[time_col].astype(str).str.strip(),
        errors='coerce',
    )


def build_receiving_by_wo(
    all_events_df: pd.DataFrame,
    receiving_event_type: str = 'Receiving In-Shop',
    event_type_col: str = 'Event Type',
    work_order_col: str = 'Work Order',
    date_col: str = 'Date',
    time_col: str = 'Time',
    customer_col_candidates: Sequence[str] = ('Company Full Name', 'Company'),
) -> pd.DataFrame:
    df = all_events_df.copy()
    df[work_order_col] = df[work_order_col].astype('string').str.strip()
    df[event_type_col] = df[event_type_col].astype('string').str.strip()

    df['EventDateTime'] = _combine_events_date_time(df, date_col=date_col, time_col=time_col)
    recv = df.loc[df[event_type_col].eq(receiving_event_type)].dropna(subset=[work_order_col, 'EventDateTime']).copy()

    customer_col = next((c for c in customer_col_candidates if c in recv.columns), None)

    receiving_by_wo = (
        recv.groupby(work_order_col, as_index=False)['EventDateTime']
            .min()
            .rename(columns={work_order_col: 'Work Order', 'EventDateTime': 'ReceivingDateTime'})
    )

    if customer_col:
        recv_sorted = recv.sort_values([work_order_col, 'EventDateTime'], ascending=[True, True])
        cust = recv_sorted.drop_duplicates(subset=[work_order_col], keep='first')[[work_order_col, customer_col]]
        cust = cust.rename(columns={work_order_col: 'Work Order', customer_col: 'Customer'})
        receiving_by_wo = receiving_by_wo.merge(cust, on='Work Order', how='left')
    else:
        receiving_by_wo['Customer'] = pd.NA

    return receiving_by_wo


# -----------------------------
# 3) Merge Events line-level with WO lines
# -----------------------------

def merge_line_level(
    all_events_df: pd.DataFrame,
    wo_lines_df: pd.DataFrame,
    work_order_col_events: str = 'Work Order',
    id_col_events: str = 'I.D.',
) -> pd.DataFrame:
    e = all_events_df.copy()
    w = wo_lines_df.copy()

    e[work_order_col_events] = e[work_order_col_events].astype('string').str.strip()
    e[id_col_events] = e[id_col_events].astype('string').str.strip()

    w['Work Order'] = w['Work Order'].astype('string').str.strip()
    w['Equipment_ID'] = w['Equipment_ID'].astype('string').str.strip()

    return e.merge(
        w,
        how='left',
        left_on=[work_order_col_events, id_col_events],
        right_on=['Work Order', 'Equipment_ID'],
        suffixes=('', '_wo'),
    )


# -----------------------------
# 4) Attach ReceivingDateTime and compute KPIs
# -----------------------------

def add_efficiency_metrics(
    line_df: pd.DataFrame,
    receiving_by_wo: pd.DataFrame,
    total_time_col: str = 'Total Time',
) -> pd.DataFrame:
    df = line_df.copy()
    df = df.merge(receiving_by_wo[['Work Order', 'ReceivingDateTime', 'Customer']], on='Work Order', how='left')

    df['EffortHours'] = pd.to_numeric(df.get(total_time_col), errors='coerce')
    if 'Actual_Time' in df.columns:
        df.loc[df['EffortHours'].isna(), 'EffortHours'] = pd.to_numeric(df.loc[df['EffortHours'].isna(), 'Actual_Time'], errors='coerce')

    df['CompletedDateTime'] = pd.to_datetime(df.get('CompletedDateTime'), errors='coerce')
    df['ReceivingDateTime'] = pd.to_datetime(df.get('ReceivingDateTime'), errors='coerce')

    df['TAT_hours'] = (df['CompletedDateTime'] - df['ReceivingDateTime']).dt.total_seconds() / 3600.0
    df['Start_Est'] = df['CompletedDateTime'] - pd.to_timedelta(df['EffortHours'], unit='h')
    df['Queue_hours'] = (df['Start_Est'] - df['ReceivingDateTime']).dt.total_seconds() / 3600.0

    return df


# -----------------------------
# 5) Work Order outcome metrics + tech list (cache-safe)
# -----------------------------

def compute_wo_metrics_and_techs(
    wo_lines_df: pd.DataFrame,
    backordered_label: str = 'Backordered',
) -> pd.DataFrame:
    """Compute WO rollups.

    IMPORTANT: WO_Techs_CompleteOnly is a pipe-delimited string for cache-safety.
    """

    df = wo_lines_df.copy()
    for col in ['Work Order', 'Equipment_ID', 'Status', 'PerformedBy_Name']:
        if col in df.columns:
            df[col] = df[col].astype('string').str.strip()

    priority = {'Complete': 3, 'Skipped': 2, backordered_label: 1}
    df['_rank'] = df['Status'].map(priority).fillna(0).astype(int)
    df = (
        df.sort_values(['Work Order', 'Equipment_ID', '_rank'], ascending=[True, True, False])
          .drop_duplicates(subset=['Work Order', 'Equipment_ID'], keep='first')
          .drop(columns=['_rank'])
    )

    g = df.groupby('Work Order', dropna=False)
    out = g.agg(
        Total_Items=('Equipment_ID', 'count'),
        Completed_Items=('Status', lambda s: (s == 'Complete').sum()),
        Skipped_Items=('Status', lambda s: (s == 'Skipped').sum()),
        Backordered_Items=('Status', lambda s: (s == backordered_label).sum()),
    ).reset_index()

    out['Incomplete_Items'] = out['Total_Items'] - out['Completed_Items'] - out['Skipped_Items'] - out['Backordered_Items']

    out['WO_Complete_%'] = out['Completed_Items'] / out['Total_Items']
    out['WO_Skipped_%'] = out['Skipped_Items'] / out['Total_Items']
    out['WO_Backordered_%'] = out['Backordered_Items'] / out['Total_Items']
    out['WO_Incomplete_%'] = out['Incomplete_Items'] / out['Total_Items']
    out['WO_Closed_%'] = (out['Completed_Items'] + out['Skipped_Items'] + out['Backordered_Items']) / out['Total_Items']

    out['WO_NonSkipped_Count'] = out['Total_Items'] - out['Skipped_Items']
    out['WO_NonSkipped_%'] = out['WO_NonSkipped_Count'] / out['Total_Items']

    complete_only = df.loc[(df['Status'] == 'Complete') & df['PerformedBy_Name'].notna() & (df['PerformedBy_Name'] != '')]
    techs_by_wo = (
        complete_only.groupby('Work Order')['PerformedBy_Name']
        .apply(lambda s: ' | '.join(sorted(set(s.tolist()))))
        .reset_index()
        .rename(columns={'PerformedBy_Name': 'WO_Techs_CompleteOnly'})
    )

    out = out.merge(techs_by_wo, on='Work Order', how='left')
    out['WO_Techs_CompleteOnly'] = out['WO_Techs_CompleteOnly'].fillna('').astype('string')

    return out


# -----------------------------
# 6) Tech association (high skip) + drilldown
# -----------------------------

def _split_techs(techs_str: Any) -> List[str]:
    if techs_str is None or (isinstance(techs_str, float) and pd.isna(techs_str)):
        return []
    s = str(techs_str).strip()
    if not s:
        return []
    return [t.strip() for t in s.split(' | ') if t.strip()]


def build_tech_wo_association(wo_rollup_df: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for _, r in wo_rollup_df.iterrows():
        techs = _split_techs(r.get('WO_Techs_CompleteOnly', ''))
        for tech in techs:
            rows.append({
                'Tech': tech,
                'Work Order': r.get('Work Order'),
                'WO_Skipped_%': r.get('WO_Skipped_%'),
                'WO_Backordered_%': r.get('WO_Backordered_%'),
                'Total_Items': r.get('Total_Items'),
                'Skipped_Items': r.get('Skipped_Items'),
                'Backordered_Items': r.get('Backordered_Items'),
            })
    return pd.DataFrame(rows)


def compute_tech_high_skip_association(
    wo_rollup_df: pd.DataFrame,
    receiving_by_wo: Optional[pd.DataFrame] = None,
    threshold: float = 0.30,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    assoc = build_tech_wo_association(wo_rollup_df)
    if assoc.empty:
        return assoc, assoc

    assoc['HighSkipWO'] = assoc['WO_Skipped_%'] >= threshold

    summary = assoc.groupby('Tech').agg(
        WO_Touched=('Work Order', 'nunique'),
        HighSkipWO_Count=('HighSkipWO', 'sum'),
        AvgSkipPct=('WO_Skipped_%', 'mean'),
        AvgBackorderedPct=('WO_Backordered_%', 'mean'),
    ).reset_index()

    summary['HighSkipWO_Rate'] = summary['HighSkipWO_Count'] / summary['WO_Touched']
    summary = summary.sort_values(['HighSkipWO_Rate', 'AvgSkipPct', 'WO_Touched'], ascending=[False, False, False])

    drill = assoc.copy()
    if receiving_by_wo is not None and not receiving_by_wo.empty:
        drill = drill.merge(receiving_by_wo[['Work Order', 'ReceivingDateTime', 'Customer']], on='Work Order', how='left')
        drill['ReceivingDate'] = pd.to_datetime(drill['ReceivingDateTime'], errors='coerce').dt.date

    return summary, drill


def top_high_skip_wos_for_tech(
    tech_drilldown_df: pd.DataFrame,
    tech_name: str,
    top_n: int = 15,
    threshold: float = 0.30,
) -> pd.DataFrame:
    df = tech_drilldown_df.copy()
    df = df.loc[df['Tech'] == tech_name]
    df = df.loc[df['WO_Skipped_%'] >= threshold]
    df = df.sort_values(['WO_Skipped_%', 'ReceivingDateTime'], ascending=[False, False])
    cols = [c for c in ['Work Order','Customer','ReceivingDateTime','WO_Skipped_%','Skipped_Items','Total_Items','Backordered_Items'] if c in df.columns]
    return df[cols].head(top_n)


# -----------------------------
# 7) Filter to Field Cal only
# -----------------------------

def filter_field_cal_only(wo_lines_df: pd.DataFrame, field_label: str = 'FIELD CALIBRATION') -> pd.DataFrame:
    if 'Related Event Type' not in wo_lines_df.columns:
        return wo_lines_df.iloc[0:0].copy()
    df = wo_lines_df.copy()
    df['Related Event Type'] = df['Related Event Type'].astype('string').str.strip()
    return df.loc[df['Related Event Type'].eq(field_label)].copy()


# -----------------------------
# 8) Skipped % trend over time per customer
# -----------------------------

def compute_skip_trend_per_customer(
    wo_rollup_df: pd.DataFrame,
    receiving_by_wo: pd.DataFrame,
    freq: str = 'M',
    min_wos_per_bucket: int = 1,
    wo_lines_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Customer outcome trend.

    - If wo_lines_df is provided, trend is computed from item-level outcomes using EffectiveDateTime.
      (Field work often has no receiving; this avoids filtering everything out.)
    - Percentages are weighted from summed item counts.

    Output columns match the existing Streamlit panel:
      Customer, Period, WO_Count, AvgSkippedPct, AvgBackorderedPct, AvgCompletePct
    """

    # Item-level path
    if wo_lines_df is not None and isinstance(wo_lines_df, pd.DataFrame) and not wo_lines_df.empty:
        df = wo_lines_df.copy()

        if 'EffectiveDateTime' not in df.columns:
            try:
                df = add_effective_date_columns(df)
            except Exception:
                pass

        if 'Company' in df.columns:
            df['Customer'] = df['Company'].astype('string').str.strip()
        elif 'WO_Company' in df.columns:
            df['Customer'] = df['WO_Company'].astype('string').str.strip()
        else:
            df['Customer'] = pd.NA

        df['Status'] = df.get('Status', pd.Series(index=df.index, dtype='string')).astype('string').str.strip().str.title()
        df = df[df['Status'].isin(['Backordered','Complete','Incomplete','Skipped'])].copy()

        df['PeriodDateTime'] = pd.to_datetime(df.get('EffectiveDateTime', pd.NaT), errors='coerce')
        df = df.dropna(subset=['Customer','PeriodDateTime']).copy()
        if df.empty:
            return pd.DataFrame(columns=['Customer','Period','WO_Count','AvgSkippedPct','AvgBackorderedPct','AvgCompletePct'])

        if freq.upper() == 'M':
            df['Period'] = df['PeriodDateTime'].dt.to_period('M').apply(lambda p: p.start_time)
        elif freq.upper() == 'W':
            df['Period'] = df['PeriodDateTime'].dt.to_period('W').apply(lambda p: p.start_time)
        else:
            raise ValueError("freq must be 'M' or 'W'")

        out = df.groupby(['Customer','Period'], as_index=False).agg(
            WO_Count=('Work Order','nunique') if 'Work Order' in df.columns else ('Customer','size'),
            Total_Items=('Equipment_ID','count') if 'Equipment_ID' in df.columns else ('Status','size'),
            Skipped=('Status', lambda s: (s == 'Skipped').sum()),
            Backordered=('Status', lambda s: (s == 'Backordered').sum()),
            Complete=('Status', lambda s: (s == 'Complete').sum()),
        )

        out['AvgSkippedPct'] = (out['Skipped'] / out['Total_Items']).where(out['Total_Items'] > 0)
        out['AvgBackorderedPct'] = (out['Backordered'] / out['Total_Items']).where(out['Total_Items'] > 0)
        out['AvgCompletePct'] = (out['Complete'] / out['Total_Items']).where(out['Total_Items'] > 0)

        out = out.loc[out['WO_Count'] >= int(min_wos_per_bucket)].copy()
        return out.sort_values(['Customer','Period'])

    # Rollup fallback (receiving-based)
    df = wo_rollup_df.merge(receiving_by_wo[['Work Order','ReceivingDateTime','Customer']], on='Work Order', how='left')
    df['ReceivingDateTime'] = pd.to_datetime(df['ReceivingDateTime'], errors='coerce')
    df = df.dropna(subset=['ReceivingDateTime','Customer'])

    if freq.upper() == 'M':
        df['Period'] = df['ReceivingDateTime'].dt.to_period('M').apply(lambda p: p.start_time)
    elif freq.upper() == 'W':
        df['Period'] = df['ReceivingDateTime'].dt.to_period('W').apply(lambda p: p.start_time)
    else:
        raise ValueError("freq must be 'M' or 'W'")

    trend = df.groupby(['Customer','Period']).agg(
        WO_Count=('Work Order','nunique'),
        Total_Items=('Total_Items','sum'),
        Skipped=('Skipped_Items','sum'),
        Backordered=('Backordered_Items','sum'),
        Complete=('Completed_Items','sum'),
    ).reset_index()

    trend['AvgSkippedPct'] = (trend['Skipped'] / trend['Total_Items']).where(trend['Total_Items'] > 0)
    trend['AvgBackorderedPct'] = (trend['Backordered'] / trend['Total_Items']).where(trend['Total_Items'] > 0)
    trend['AvgCompletePct'] = (trend['Complete'] / trend['Total_Items']).where(trend['Total_Items'] > 0)

    return trend.loc[trend['WO_Count'] >= int(min_wos_per_bucket)].copy().sort_values(['Customer','Period'])

def build_efficiency_datasets(
    all_events_df: pd.DataFrame,
    all_wos_export_path: str | Path,
    names_numbers_path: str | Path,
    skip_threshold: float = 0.30,
) -> Dict[str, pd.DataFrame]:
    wo_lines = parse_all_wos_subtables(all_wos_export_path, names_numbers_path)
    wo_lines_field = filter_field_cal_only(wo_lines)

    receiving_by_wo = build_receiving_by_wo(all_events_df)

    events_merged_line = merge_line_level(all_events_df, wo_lines)
    events_merged_kpis = add_efficiency_metrics(events_merged_line, receiving_by_wo)

    wo_rollup = compute_wo_metrics_and_techs(wo_lines)
    wo_rollup_field = compute_wo_metrics_and_techs(wo_lines_field) if not wo_lines_field.empty else wo_rollup.iloc[0:0].copy()

    tech_summary, tech_drilldown = compute_tech_high_skip_association(wo_rollup, receiving_by_wo, threshold=skip_threshold)
    tech_summary_field, tech_drilldown_field = compute_tech_high_skip_association(wo_rollup_field, receiving_by_wo, threshold=skip_threshold)

    return {
        'wo_lines': wo_lines,
        'wo_lines_field': wo_lines_field,
        'receiving_by_wo': receiving_by_wo,
        'events_merged_line': events_merged_line,
        'events_merged_kpis': events_merged_kpis,
        'wo_rollup': wo_rollup,
        'wo_rollup_field': wo_rollup_field,
        'tech_summary': tech_summary,
        'tech_drilldown': tech_drilldown,
        'tech_summary_field': tech_summary_field,
        'tech_drilldown_field': tech_drilldown_field,
    }


# -----------------------------
# Optional Streamlit helpers
# -----------------------------

def streamlit_customer_skip_trend_panel(customer_trend_monthly_df: pd.DataFrame, default_customer: Optional[str] = None) -> None:
    import streamlit as st
    import altair as alt

    st.subheader('Customer Outcome Trend (Monthly)')
    customers = sorted(customer_trend_monthly_df['Customer'].dropna().unique().tolist())
    if not customers:
        st.info('No customer trend data available.')
        return

    if default_customer is None or default_customer not in customers:
        default_customer = customers[0]

    cust = st.selectbox('Customer', customers, index=customers.index(default_customer))
    df = customer_trend_monthly_df.loc[customer_trend_monthly_df['Customer'] == cust].copy()
    df['Period'] = pd.to_datetime(df['Period'], errors='coerce')

    long_df = df.melt(
        id_vars=['Customer','Period','WO_Count'],
        value_vars=['AvgSkippedPct','AvgBackorderedPct','AvgCompletePct'],
        var_name='Metric',
        value_name='Value'
    )

    metric_labels = {
        'AvgSkippedPct': 'Skipped %',
        'AvgBackorderedPct': 'Backordered %',
        'AvgCompletePct': 'Complete %',
    }
    long_df['Metric'] = long_df['Metric'].map(metric_labels).fillna(long_df['Metric'])

    line = alt.Chart(long_df).mark_line(point=True).encode(
        x=alt.X('Period:T', title='Month'),
        y=alt.Y('Value:Q', title='Percent', axis=alt.Axis(format='%'), scale=alt.Scale(domain=[0,1])),
        color=alt.Color('Metric:N', title='Metric'),
        tooltip=[
            alt.Tooltip('Period:T', title='Month'),
            alt.Tooltip('Metric:N'),
            alt.Tooltip('Value:Q', format='.1%'),
            alt.Tooltip('WO_Count:Q', title='# WOs'),
        ]
    ).properties(height=320)

    bars = alt.Chart(df).mark_bar(opacity=0.12).encode(
        x=alt.X('Period:T'),
        y=alt.Y('WO_Count:Q', title='# WOs'),
        tooltip=[alt.Tooltip('Period:T', title='Month'), alt.Tooltip('WO_Count:Q', title='# WOs')]
    ).properties(height=120)

    st.altair_chart(alt.vconcat(line, bars), use_container_width=True)


def streamlit_tech_drilldown_panel(
    tech_summary_df: pd.DataFrame,
    tech_drilldown_df: pd.DataFrame,
    threshold: float = 0.30,
    top_n: int = 15,
) -> None:
    import streamlit as st

    if tech_summary_df is None or tech_summary_df.empty:
        st.info('No tech association data available.')
        return

    st.caption(f"High-skip threshold: {threshold:.0%}. Techs derived from COMPLETE rows only.")
    st.dataframe(tech_summary_df, use_container_width=True)

    techs = sorted(tech_summary_df['Tech'].dropna().unique().tolist())
    if not techs:
        return

    tech = st.selectbox('Select tech', techs)
    drill = top_high_skip_wos_for_tech(tech_drilldown_df, tech, top_n=top_n, threshold=threshold)
    st.markdown('**Top high-skip work orders**')
    st.dataframe(drill, use_container_width=True)

    
import pandas as pd

def add_effective_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Computed helper for filtering/trending (computed fields).
    out = df.copy()
    out['CompletedDateTime'] = pd.to_datetime(out.get('CompletedDateTime'), errors='coerce')
    out['WO_CompletedDateTime'] = pd.to_datetime(out.get('WO_CompletedDateTime'), errors='coerce')
    status = out.get('Status', pd.Series(index=out.index, dtype='string')).astype('string').str.strip().str.title()

    out['EffectiveDateTime'] = pd.NaT
    m_complete = status.eq('Complete')
    out.loc[m_complete, 'EffectiveDateTime'] = out.loc[m_complete, 'CompletedDateTime']
    out.loc[~m_complete, 'EffectiveDateTime'] = out.loc[~m_complete, 'WO_CompletedDateTime']

    if 'Work Order' in out.columns:
        proxy = (out.loc[m_complete, ['Work Order','CompletedDateTime']]
                 .dropna(subset=['Work Order','CompletedDateTime'])
                 .groupby('Work Order', as_index=False)['CompletedDateTime']
                 .max()
                 .rename(columns={'CompletedDateTime':'WO_Completed_Proxy'}))
        if not proxy.empty:
            out = out.merge(proxy, on='Work Order', how='left')
            out['EffectiveDateTime'] = out['EffectiveDateTime'].fillna(out['WO_Completed_Proxy'])

    out['EffectiveMonth'] = pd.to_datetime(out['EffectiveDateTime'], errors='coerce').dt.to_period('M').dt.start_time
    return out




# Compatibility alias
add_effictive_date_columns = add_effective_date_columns
